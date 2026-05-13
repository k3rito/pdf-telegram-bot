"""
بوت تيليجرام لخدمات PDF
يدعم: دمج، تقسيم، ضغط، استخراج نصوص/صور، تحويل صور→PDF، تحويل PDF→Word/Excel
"""

import os
import io
import zipfile
import logging
import tempfile
from pathlib import Path
from dotenv import load_dotenv

load_dotenv()

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    CallbackQueryHandler, ContextTypes, ConversationHandler, filters
)

# ─── إعدادات ───────────────────────────────────────────────────────────────────
BOT_TOKEN = os.getenv("BOT_TOKEN")
if not BOT_TOKEN:
    raise ValueError("❌ لم يتم تعيين BOT_TOKEN في ملف .env")

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# حالات المحادثة
CHOOSING, COLLECTING_PDFS, COLLECTING_IMAGES, WAITING_SPLIT_RANGE = range(4)

# ─── لوحة المفاتيح الرئيسية ────────────────────────────────────────────────────
MAIN_MENU = InlineKeyboardMarkup([
    [InlineKeyboardButton("📎 دمج ملفات PDF",        callback_data="merge")],
    [InlineKeyboardButton("✂️ تقسيم PDF",            callback_data="split")],
    [InlineKeyboardButton("🗜️ ضغط PDF",              callback_data="compress")],
    [InlineKeyboardButton("📝 استخراج النصوص",       callback_data="extract_text")],
    [InlineKeyboardButton("🖼️ استخراج الصور",        callback_data="extract_images")],
    [InlineKeyboardButton("🖼️➡️📄 صور إلى PDF",     callback_data="images_to_pdf")],
    [InlineKeyboardButton("📄➡️📝 PDF إلى Word",     callback_data="pdf_to_word")],
    [InlineKeyboardButton("📊 PDF إلى Excel",         callback_data="pdf_to_excel")],
])

CANCEL_BTN = InlineKeyboardMarkup([
    [InlineKeyboardButton("❌ إلغاء", callback_data="cancel")]
])


# ─── /start ────────────────────────────────────────────────────────────────────
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text(
        "👋 *مرحباً بك في بوت PDF!*\n\nاختر الخدمة التي تريدها:",
        reply_markup=MAIN_MENU,
        parse_mode="Markdown",
    )
    return CHOOSING


# ─── اختيار الخدمة ─────────────────────────────────────────────────────────────
async def choose_service(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    service = query.data

    if service == "cancel":
        return await cancel(update, context)

    context.user_data["service"] = service
    context.user_data["files"] = []

    messages = {
        "merge":          "📎 *دمج PDF*\nأرسل ملفات PDF واحداً تلو الآخر، ثم اكتب /done عند الانتهاء.",
        "split":          "✂️ *تقسيم PDF*\nأرسل ملف PDF واحد.",
        "compress":       "🗜️ *ضغط PDF*\nأرسل ملف PDF واحد.",
        "extract_text":   "📝 *استخراج النصوص*\nأرسل ملف PDF واحد.",
        "extract_images": "🖼️ *استخراج الصور*\nأرسل ملف PDF واحد.",
        "images_to_pdf":  "🖼️➡️📄 *صور إلى PDF*\nأرسل الصور واحدة تلو الأخرى، ثم اكتب /done.",
        "pdf_to_word":    "📄➡️📝 *PDF إلى Word*\nأرسل ملف PDF واحد.",
        "pdf_to_excel":   "📊 *PDF إلى Excel*\nأرسل ملف PDF واحد.",
    }

    state = COLLECTING_IMAGES if service == "images_to_pdf" else COLLECTING_PDFS

    await query.edit_message_text(
        messages[service],
        reply_markup=CANCEL_BTN,
        parse_mode="Markdown",
    )
    return state


# ─── استقبال ملفات PDF ──────────────────────────────────────────────────────────
async def receive_pdf(update: Update, context: ContextTypes.DEFAULT_TYPE):
    service = context.user_data.get("service")
    doc = update.message.document

    if not doc or not doc.file_name.lower().endswith(".pdf"):
        await update.message.reply_text("❌ الرجاء إرسال ملف PDF فقط.")
        return COLLECTING_PDFS

    file = await doc.get_file()
    data = await file.download_as_bytearray()
    context.user_data["files"].append((doc.file_name, bytes(data)))

    # خدمات تحتاج ملف واحد فقط → نعالج مباشرة
    single_file_services = {"split", "compress", "extract_text", "extract_images", "pdf_to_word", "pdf_to_excel"}
    if service in single_file_services:
        await update.message.reply_text("⏳ جاري المعالجة...")
        return await process_service(update, context)

    count = len(context.user_data["files"])
    await update.message.reply_text(
        f"✅ تم استلام الملف ({count}). أرسل ملف آخر أو اكتب /done للدمج.",
        reply_markup=CANCEL_BTN,
    )
    return COLLECTING_PDFS


# ─── استقبال الصور ─────────────────────────────────────────────────────────────
async def receive_image(update: Update, context: ContextTypes.DEFAULT_TYPE):
    photo = update.message.photo or (
        [update.message.document] if update.message.document and
        update.message.document.mime_type.startswith("image/") else None
    )

    if not photo:
        await update.message.reply_text("❌ الرجاء إرسال صورة.")
        return COLLECTING_IMAGES

    file_obj = photo[-1] if isinstance(photo[0], type(update.message.photo[0] if update.message.photo else None)) else photo[0]
    file = await file_obj.get_file()
    data = await file.download_as_bytearray()
    context.user_data["files"].append(("image", bytes(data)))

    count = len(context.user_data["files"])
    await update.message.reply_text(
        f"✅ تم استلام الصورة ({count}). أرسل صورة أخرى أو اكتب /done.",
        reply_markup=CANCEL_BTN,
    )
    return COLLECTING_IMAGES


# ─── /done → معالجة ─────────────────────────────────────────────────────────────
async def done(update: Update, context: ContextTypes.DEFAULT_TYPE):
    files = context.user_data.get("files", [])
    service = context.user_data.get("service")

    if not files:
        await update.message.reply_text("❌ لم يتم إرسال أي ملف.")
        return CHOOSING

    if service == "merge" and len(files) < 2:
        await update.message.reply_text("❌ الدمج يحتاج ملفَين على الأقل.")
        return COLLECTING_PDFS

    await update.message.reply_text("⏳ جاري المعالجة...")
    return await process_service(update, context)


# ─── معالجة الخدمة ─────────────────────────────────────────────────────────────
async def process_service(update: Update, context: ContextTypes.DEFAULT_TYPE):
    service = context.user_data["service"]
    files   = context.user_data["files"]
    msg     = update.message or update.callback_query.message

    try:
        if service == "merge":
            result, name = merge_pdfs(files)
            await msg.reply_document(document=result, filename=name, caption="✅ تم دمج الملفات!")

        elif service == "split":
            results = split_pdf(files[0][1])
            zip_buf = make_zip(results)
            await msg.reply_document(document=zip_buf, filename="pages.zip", caption=f"✅ تم تقسيم PDF إلى {len(results)} صفحة!")

        elif service == "compress":
            result, orig, comp = compress_pdf(files[0][1])
            saved = round((1 - comp / orig) * 100)
            await msg.reply_document(document=result, filename="compressed.pdf",
                caption=f"✅ تم الضغط!\n📦 الأصلي: {orig//1024} كيلو\n📦 المضغوط: {comp//1024} كيلو\n💾 وفّرنا: {saved}%")

        elif service == "extract_text":
            text = extract_text(files[0][1])
            if len(text) > 4000:
                buf = io.BytesIO(text.encode("utf-8"))
                buf.name = "extracted_text.txt"
                await msg.reply_document(document=buf, filename="extracted_text.txt", caption="✅ تم استخراج النصوص!")
            else:
                await msg.reply_text(f"✅ النصوص المستخرجة:\n\n{text}")

        elif service == "extract_images":
            images = extract_images_from_pdf(files[0][1])
            if not images:
                await msg.reply_text("⚠️ لم يتم العثور على صور في هذا الملف.")
            else:
                zip_buf = make_zip(images)
                await msg.reply_document(document=zip_buf, filename="images.zip",
                    caption=f"✅ تم استخراج {len(images)} صورة!")

        elif service == "images_to_pdf":
            result = images_to_pdf(files)
            await msg.reply_document(document=result, filename="output.pdf", caption="✅ تم تحويل الصور إلى PDF!")

        elif service == "pdf_to_word":
            result = pdf_to_word(files[0][1])
            await msg.reply_document(document=result, filename="output.docx", caption="✅ تم التحويل إلى Word!")

        elif service == "pdf_to_excel":
            result = pdf_to_excel(files[0][1])
            await msg.reply_document(document=result, filename="output.xlsx", caption="✅ تم التحويل إلى Excel!")

    except Exception as e:
        logger.exception("خطأ في المعالجة")
        await msg.reply_text(f"❌ حدث خطأ: {e}")

    # العودة للقائمة
    await msg.reply_text("🔁 اختر خدمة أخرى:", reply_markup=MAIN_MENU)
    context.user_data.clear()
    return CHOOSING


# ─── وظائف PDF ─────────────────────────────────────────────────────────────────

def merge_pdfs(files):
    from pypdf import PdfWriter, PdfReader
    writer = PdfWriter()
    for name, data in files:
        reader = PdfReader(io.BytesIO(data))
        for page in reader.pages:
            writer.add_page(page)
    buf = io.BytesIO()
    writer.write(buf)
    buf.seek(0)
    return buf, "merged.pdf"


def split_pdf(data):
    from pypdf import PdfReader, PdfWriter
    reader = PdfReader(io.BytesIO(data))
    results = []
    for i, page in enumerate(reader.pages):
        writer = PdfWriter()
        writer.add_page(page)
        buf = io.BytesIO()
        writer.write(buf)
        buf.seek(0)
        results.append((f"page_{i+1}.pdf", buf.read()))
    return results


def compress_pdf(data):
    from pypdf import PdfReader, PdfWriter
    reader = PdfReader(io.BytesIO(data))
    writer = PdfWriter()
    for page in reader.pages:
        page.compress_content_streams()
        writer.add_page(page)
    writer.add_metadata(reader.metadata or {})
    buf = io.BytesIO()
    writer.write(buf)
    buf.seek(0)
    compressed = buf.read()
    buf.seek(0)
    return buf, len(data), len(compressed)


def extract_text(data):
    import pdfplumber
    text = ""
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for i, page in enumerate(pdf.pages):
            page_text = page.extract_text() or ""
            text += f"── صفحة {i+1} ──\n{page_text}\n\n"
    return text.strip() or "لم يتم العثور على نصوص."


def extract_images_from_pdf(data):
    from pypdf import PdfReader
    reader = PdfReader(io.BytesIO(data))
    images = []
    count = 0
    for page in reader.pages:
        for img in page.images:
            count += 1
            images.append((f"image_{count}.{img.name.split('.')[-1]}", img.data))
    return images


def images_to_pdf(files):
    from PIL import Image
    pdf_images = []
    for _, data in files:
        img = Image.open(io.BytesIO(data)).convert("RGB")
        pdf_images.append(img)
    buf = io.BytesIO()
    if pdf_images:
        pdf_images[0].save(buf, format="PDF", save_all=True, append_images=pdf_images[1:])
    buf.seek(0)
    return buf


def pdf_to_word(data):
    from pdf2docx import Converter
    import tempfile, os
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp_pdf:
        tmp_pdf.write(data)
        tmp_pdf_path = tmp_pdf.name
    tmp_docx_path = tmp_pdf_path.replace(".pdf", ".docx")
    try:
        cv = Converter(tmp_pdf_path)
        cv.convert(tmp_docx_path, start=0, end=None)
        cv.close()
        with open(tmp_docx_path, "rb") as f:
            docx_data = f.read()
    finally:
        os.unlink(tmp_pdf_path)
        if os.path.exists(tmp_docx_path):
            os.unlink(tmp_docx_path)
    return io.BytesIO(docx_data)


def pdf_to_excel(data):
    import pdfplumber
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for i, page in enumerate(pdf.pages):
            tables = page.extract_tables()
            if tables:
                for j, table in enumerate(tables):
                    ws = wb.create_sheet(title=f"P{i+1}_T{j+1}"[:31])
                    for row in table:
                        ws.append([cell or "" for cell in row])
            else:
                # نص عادي في حالة ما في جداول
                ws = wb.create_sheet(title=f"Page_{i+1}"[:31])
                text = page.extract_text() or ""
                for k, line in enumerate(text.split("\n"), 1):
                    ws.cell(row=k, column=1, value=line)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def make_zip(files):
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, data in files:
            zf.writestr(name, data)
    buf.seek(0)
    return buf


# ─── إلغاء ─────────────────────────────────────────────────────────────────────
async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    msg = update.message or update.callback_query.message
    await msg.reply_text("❌ تم الإلغاء. اختر خدمة:", reply_markup=MAIN_MENU)
    return CHOOSING


# ─── تشغيل البوت ───────────────────────────────────────────────────────────────
def main():
    app = Application.builder().token(BOT_TOKEN).build()

    conv = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            CHOOSING: [
                CallbackQueryHandler(choose_service),
            ],
            COLLECTING_PDFS: [
                MessageHandler(filters.Document.ALL | filters.PDF, receive_pdf),
                CommandHandler("done", done),
                CallbackQueryHandler(cancel, pattern="cancel"),
            ],
            COLLECTING_IMAGES: [
                MessageHandler(filters.PHOTO | filters.Document.IMAGE, receive_image),
                CommandHandler("done", done),
                CallbackQueryHandler(cancel, pattern="cancel"),
            ],
        },
        fallbacks=[
            CommandHandler("start", start),
            CommandHandler("cancel", cancel),
        ],
    )

    app.add_handler(conv)
    print("🤖 البوت يعمل...")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
