from __future__ import annotations

import zipfile
from pathlib import Path

from services.types import ServiceResult


def images_to_pdf(paths: list[Path], output_dir: Path) -> ServiceResult:
    from PIL import Image

    output_path = output_dir / "output.pdf"
    pdf_images = []
    for path in paths:
        img = Image.open(path).convert("RGB")
        pdf_images.append(img)

    if pdf_images:
        pdf_images[0].save(output_path, format="PDF", save_all=True, append_images=pdf_images[1:])

    return ServiceResult(
        kind="document",
        path=output_path,
        filename="output.pdf",
        caption="\u2705 \u062a\u0645 \u062a\u062d\u0648\u064a\u0644 \u0627\u0644\u0635\u0648\u0631 \u0625\u0644\u0649 PDF!",
    )


def pdf_to_word(path: Path, output_dir: Path) -> ServiceResult:
    from pdf2docx import Converter

    output_path = output_dir / "output.docx"
    converter = Converter(str(path))
    converter.convert(str(output_path), start=0, end=None)
    converter.close()

    return ServiceResult(
        kind="document",
        path=output_path,
        filename="output.docx",
        caption="\u2705 \u062a\u0645 \u0627\u0644\u062a\u062d\u0648\u064a\u0644 \u0625\u0644\u0649 Word!",
    )


def pdf_to_excel(path: Path, output_dir: Path) -> ServiceResult:
    import pdfplumber
    import openpyxl

    output_path = output_dir / "output.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    with pdfplumber.open(path) as pdf:
        for page_index, page in enumerate(pdf.pages):
            tables = page.extract_tables()
            if tables:
                for table_index, table in enumerate(tables):
                    ws = wb.create_sheet(title=f"P{page_index + 1}_T{table_index + 1}"[:31])
                    for row in table:
                        ws.append([cell or "" for cell in row])
            else:
                ws = wb.create_sheet(title=f"Page_{page_index + 1}"[:31])
                text = page.extract_text() or ""
                for line_index, line in enumerate(text.split("\n"), 1):
                    ws.cell(row=line_index, column=1, value=line)

    wb.save(output_path)

    return ServiceResult(
        kind="document",
        path=output_path,
        filename="output.xlsx",
        caption="\u2705 \u062a\u0645 \u0627\u0644\u062a\u062d\u0648\u064a\u0644 \u0625\u0644\u0649 Excel!",
    )


def pdf_to_images(path: Path, output_dir: Path) -> ServiceResult:
    import fitz

    doc = fitz.open(path)
    page_count = doc.page_count
    zip_path = output_dir / "pages_images.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for index in range(page_count):
            page = doc.load_page(index)
            pix = page.get_pixmap(dpi=200)
            img_path = output_dir / f"page_{index + 1}.png"
            pix.save(str(img_path))
            zf.write(img_path, arcname=img_path.name)
            img_path.unlink(missing_ok=True)
    doc.close()

    return ServiceResult(
        kind="document",
        path=zip_path,
        filename="pages_images.zip",
        caption=f"\u2705 \u062a\u0645 \u062a\u062d\u0648\u064a\u0644 {page_count} \u0635\u0641\u062d\u0629 \u0625\u0644\u0649 \u0635\u0648\u0631!",
    )
